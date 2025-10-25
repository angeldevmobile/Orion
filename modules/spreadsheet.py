"""
Módulo spreadsheet de Orion
Permite crear, leer y escribir hojas de cálculo (.xlsx, .csv, .orion)
sin depender de APIs externas, y preparado para sincronización Orion Cloud.
"""

import os
import csv
import json
from openpyxl import Workbook, load_workbook
from modules import net 
from modules.sheets_localbridge import register as bridge_register


class OrionSpreadsheet:
    """Controlador universal de hojas de cálculo Orion."""
    _registry = {}  # para manejar identificadores (IDs) internos

    def __init__(self, filename):
        self.filename = filename
        self.ext = os.path.splitext(filename)[1].lower()

        # --- Soporte local ---
        if self.ext == ".xlsx":
            if os.path.exists(filename):
                self.wb = load_workbook(filename)
                self.ws = self.wb.active
            else:
                self.wb = Workbook()
                self.ws = self.wb.active

        elif self.ext == ".csv":
            self.rows = []
            if os.path.exists(filename):
                with open(filename, newline='', encoding='utf-8') as f:
                    self.rows = list(csv.reader(f))
            # CORRECCIÓN: Inicializar rows vacío si el archivo no existe

        # --- Formato nativo Orion ---
        elif self.ext in [".orion", ".osheet"]:
            if os.path.exists(filename):
                with open(filename, "r", encoding="utf-8") as f:
                    self.data = json.load(f)
            else:
                self.data = {"headers": [], "rows": []}

        else:
            raise ValueError(f"Formato no soportado: {self.ext}")

    # -----------------------------------------------------------
    # Operaciones básicas
    # -----------------------------------------------------------

    def write(self, cell, value):
        if self.ext == ".xlsx":
            self.ws[cell] = value
        elif self.ext == ".csv":
            # CORRECCIÓN: Agregar soporte para escribir en CSV por celdas
            row_idx = int(''.join(filter(str.isdigit, cell))) - 1
            col_letter = ''.join(filter(str.isalpha, cell)).upper()
            col_idx = ord(col_letter) - 65  # A → 0
            
            # Asegurar que hay suficientes filas
            while len(self.rows) <= row_idx:
                self.rows.append([])
            
            # Asegurar que hay suficientes columnas en la fila
            while len(self.rows[row_idx]) <= col_idx:
                self.rows[row_idx].append("")
            
            self.rows[row_idx][col_idx] = str(value)
            
        elif self.ext in [".orion", ".osheet"]:
            # escritura simbólica tipo A1, B2, etc.
            row_idx = int(''.join(filter(str.isdigit, cell))) - 1
            col_letter = ''.join(filter(str.isalpha, cell)).upper()
            col_idx = ord(col_letter) - 65  # A → 0
            while len(self.data["rows"]) <= row_idx:
                self.data["rows"].append([])
            row = self.data["rows"][row_idx]
            while len(row) <= col_idx:
                row.append("")
            row[col_idx] = value
        else:
            raise TypeError("La escritura directa solo aplica a .xlsx, .csv o .orion")

    def append(self, values):
        """Agrega una fila completa al final de la hoja."""
        if self.ext == ".xlsx":
            self.ws.append(values)
        elif self.ext == ".csv":
            # CORRECCIÓN: Asegurar que rows está inicializado
            if not hasattr(self, 'rows'):
                self.rows = []
            self.rows.append([str(v) for v in values])  # Convertir a strings para CSV
        elif self.ext in [".orion", ".osheet"]:
            self.data["rows"].append(values)
        else:
            raise TypeError(f"Método append no soportado para formato {self.ext}")

    def read(self, cell):
        if self.ext == ".xlsx":
            return self.ws[cell].value
        elif self.ext == ".csv":
            # CORRECCIÓN: Agregar soporte para leer CSV por celdas
            row_idx = int(''.join(filter(str.isdigit, cell))) - 1
            col_letter = ''.join(filter(str.isalpha, cell)).upper()
            col_idx = ord(col_letter) - 65
            try:
                return self.rows[row_idx][col_idx]
            except (IndexError, AttributeError):
                return None
        elif self.ext in [".orion", ".osheet"]:
            row_idx = int(''.join(filter(str.isdigit, cell))) - 1
            col_letter = ''.join(filter(str.isalpha, cell)).upper()
            col_idx = ord(col_letter) - 65
            try:
                return self.data["rows"][row_idx][col_idx]
            except IndexError:
                return None
        else:
            raise TypeError("Lectura no soportada para este formato")

    def save(self):
        if self.ext == ".xlsx":
            self.wb.save(self.filename)
        elif self.ext == ".csv":
            # CORRECCIÓN: Asegurar que el directorio existe
            os.makedirs(os.path.dirname(self.filename) if os.path.dirname(self.filename) else ".", exist_ok=True)
            with open(self.filename, "w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                if hasattr(self, 'rows') and self.rows:
                    writer.writerows(self.rows)
        elif self.ext in [".orion", ".osheet"]:
            os.makedirs(os.path.dirname(self.filename) if os.path.dirname(self.filename) else ".", exist_ok=True)
            with open(self.filename, "w", encoding="utf-8") as f:
                json.dump(self.data, f, indent=2, ensure_ascii=False)
        return f"Archivo guardado: {self.filename}"

    # -----------------------------------------------------------
    # Métodos adicionales para compatibilidad
    # -----------------------------------------------------------
    
    def get_all_data(self):
        """Obtiene todos los datos de la hoja."""
        if self.ext == ".xlsx":
            return [[cell.value for cell in row] for row in self.ws.iter_rows()]
        elif self.ext == ".csv":
            return self.rows if hasattr(self, 'rows') else []
        elif self.ext in [".orion", ".osheet"]:
            return self.data["rows"]
        
    def clear(self):
        """Limpia todo el contenido de la hoja."""
        if self.ext == ".xlsx":
            self.wb = Workbook()
            self.ws = self.wb.active
        elif self.ext == ".csv":
            self.rows = []
        elif self.ext in [".orion", ".osheet"]:
            self.data = {"headers": [], "rows": []}

    # -----------------------------------------------------------
    # Sincronización Cloud
    # -----------------------------------------------------------

    def sync(self, endpoint):
        """Sincroniza la hoja con un endpoint Orion Cloud."""
        if self.ext in [".orion", ".osheet"]:
            payload = {"sheet": self.data, "source": self.filename}
        else:
            payload = {"message": f"Sync desde {self.filename}"}

        try:
            res = net.transmit(endpoint, json_data=payload)
            if res.status in [200, 201]:
                return f"Sincronización exitosa con {endpoint}"
            else:
                return f"Error de sincronización: {res.status}"
        except Exception as e:
            return f"Fallo al conectar con Orion Cloud: {e}"

    # -----------------------------------------------------------
    # Gestión de registros / IDs Orion
    # -----------------------------------------------------------

    @classmethod
    def register(cls, id_name, filename):
        cls._registry[id_name] = filename
        return f"Hoja registrada con ID '{id_name}' → {filename}"

    @classmethod
    def attach(cls, id=None, filename=None):
        if not filename and id in cls._registry:
            filename = cls._registry[id]
        if not filename:
            raise ValueError("Debe especificarse un filename o un ID válido")
        return OrionSpreadsheet(filename)


# -----------------------------------------------------------
# API expuesta a Orion
# -----------------------------------------------------------

def create(filename):
    return OrionSpreadsheet(filename)

def attach(id=None, filename=None):
    return OrionSpreadsheet.attach(id=id, filename=filename)

def register(id_name, filename):
    OrionSpreadsheet.register(id_name, filename)
    bridge_register(id_name, filename)  
    return f"Hoja registrada con ID '{id_name}' → {filename}"
