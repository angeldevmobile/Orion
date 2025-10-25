"""
Orion LocalSheet Bridge (OLSB)
────────────────────────────────────────────
Sistema de hojas de cálculo universal sin API externa.
Lee y escribe en archivos locales, y prepara sincronización
inteligente hacia Orion Cloud o nodos remotos.

Formato universal: orion://sheet/<id>/<celda>
"""

import os
import time
import threading
from openpyxl import load_workbook, Workbook
import csv

SYNC_PATH = "storage/sync"

# ============================================================
# OSyncCache Core: Protocolo Orion Sync (solo cache local)
# ============================================================

class OSyncCache:
    """Manejador del protocolo Orion Sync (solo cache local, no exporta JSON)."""
    @staticmethod
    def push(sheet_id):
        cache = os.path.join(SYNC_PATH, f"{sheet_id}.orioncache")
        if os.path.exists(cache):
            print(f"[OSyncCache] Pushing cached changes for {sheet_id} ...")
            return True
        else:
            print(f"[OSyncCache] No changes to push for {sheet_id}.")
            return False

    @staticmethod
    def pull(sheet_id):
        print(f"[OSyncCache] Pulling updates for {sheet_id} ... (not implemented yet)")
        return False

    @staticmethod
    def status(sheet_id):
        cache = os.path.join(SYNC_PATH, f"{sheet_id}.orioncache")
        if os.path.exists(cache):
            with open(cache, "r", encoding="utf-8") as f:
                lines = f.readlines()
            return {
                "sheet_id": sheet_id,
                "has_changes": len(lines) > 0,
                "last_change": lines[-1].strip() if lines else None,
                "total_changes": len(lines),
                "synced": False
            }
        else:
            return {
                "sheet_id": sheet_id,
                "has_changes": False,
                "last_change": None,
                "total_changes": 0,
                "synced": True
            }

    @staticmethod
    def list_synced():
        os.makedirs(SYNC_PATH, exist_ok=True)
        synced_sheets = []
        for file in os.listdir(SYNC_PATH):
            if file.endswith(".orioncache"):
                sheet_id = file.replace(".orioncache", "")
                status = OSyncCache.status(sheet_id)
                synced_sheets.append(status)
        return synced_sheets

    @staticmethod
    def log(sheet_id, action, cell, value):
        os.makedirs(SYNC_PATH, exist_ok=True)
        cache_file = os.path.join(SYNC_PATH, f"{sheet_id}.orioncache")
        with open(cache_file, "a", encoding="utf-8") as f:
            timestamp = time.strftime("%Y-%m-%d %H:%M:%S")
            f.write(f"[{timestamp}] {action.upper()} {cell} = {value}\n")


# ============================================================
# Bridge: conexión local a hojas de cálculo
# ============================================================

class LocalSheetBridge:
    """Puente local que permite acceso a hojas de cálculo sin API."""
    _registry = {}

    def __init__(self, sheet_id, path):
        self.sheet_id = sheet_id
        self.path = path
        self.ext = os.path.splitext(path)[1].lower()
        self._load()

    def _load(self):
        if self.ext == ".xlsx":
            if os.path.exists(self.path):
                self.wb = load_workbook(self.path)
                self.ws = self.wb.active
            else:
                self.wb = Workbook()
                self.ws = self.wb.active
        elif self.ext == ".csv":
            self.rows = []
            if os.path.exists(self.path):
                with open(self.path, newline="", encoding="utf-8") as f:
                    reader = csv.reader(f)
                    self.rows = list(reader)
        else:
            raise ValueError(f"Formato no soportado: {self.ext}")

    def read(self, cell):
        if self.ext == ".xlsx":
            return self.ws[cell].value
        elif self.ext == ".csv":
            row, col = self._parse_csv_cell(cell)
            return self.rows[row][col]

    def write(self, cell, value):
        if self.ext == ".xlsx":
            self.ws[cell] = value
        elif self.ext == ".csv":
            row, col = self._parse_csv_cell(cell)
            while len(self.rows) <= row:
                self.rows.append([])
            while len(self.rows[row]) <= col:
                self.rows[row].append("")
            self.rows[row][col] = value
        OSyncCache.log(self.sheet_id, "write", cell, value)

    def append(self, values):
        if self.ext == ".xlsx":
            self.ws.append(values)
        elif self.ext == ".csv":
            self.rows.append(values)
        OSyncCache.log(self.sheet_id, "append", "-", values)

    def save(self):
        if self.ext == ".xlsx":
            self.wb.save(self.path)
        elif self.ext == ".csv":
            with open(self.path, "w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                writer.writerows(self.rows)
        print(f"[Bridge] {self.sheet_id} saved → {self.path}")

    def _parse_csv_cell(self, cell):
        # Convierte 'A1' a (fila, columna)
        col = ord(cell[0].upper()) - 65
        row = int(cell[1:]) - 1
        return row, col

    # ============================================================
    # Registro Orion
    # ============================================================

    @classmethod
    def register(cls, id_name, path):
        cls._registry[id_name] = path
        return f"Hoja registrada con ID '{id_name}' → {path}"

    @classmethod
    def attach(cls, id_name):
        if id_name not in cls._registry:
            raise ValueError(f"ID no registrado: {id_name}")
        return LocalSheetBridge(id_name, cls._registry[id_name])


# ============================================================
# API pública Orion
# ============================================================

def register(id_name, path):
    return LocalSheetBridge.register(id_name, path)

def attach(id_name):
    return LocalSheetBridge.attach(id_name)

# NO expongas push/pull aquí, deja que el sistema use el de sync/osync.py

def status(id_name):
    return OSyncCache.status(id_name)

def list_synced():
    return OSyncCache.list_synced()

def clear_cache(id_name):
    cache_file = os.path.join(SYNC_PATH, f"{id_name}.orioncache")
    if os.path.exists(cache_file):
        os.remove(cache_file)
        return f"Cache cleared for {id_name}"
    return f"No cache found for {id_name}"

# ============================================================
# Funciones de conveniencia adicionales
# ============================================================

def create_sheet(id_name, path, data=None):
    """Crea y registra una nueva hoja de cálculo."""
    register(id_name, path)
    sheet = attach(id_name)
    
    if data:
        if isinstance(data, list):
            for row in data:
                sheet.append(row)
        elif isinstance(data, dict):
            for cell, value in data.items():
                sheet.write(cell, value)
    
    sheet.save()
    return sheet

def quick_write(id_name, cell, value):
    """Escribir rápidamente a una celda."""
    sheet = attach(id_name)
    sheet.write(cell, value)
    sheet.save()
    return f"Written {value} to {cell} in {id_name}"

def quick_read(id_name, cell):
    """Leer rápidamente de una celda."""
    sheet = attach(id_name)
    return sheet.read(cell)

def get_sheet_info(id_name):
    """Obtiene información sobre una hoja registrada."""
    if id_name in LocalSheetBridge._registry:
        path = LocalSheetBridge._registry[id_name]
        exists = os.path.exists(path)
        ext = os.path.splitext(path)[1].lower()
        
        info = {
            "id": id_name,
            "path": path,
            "exists": exists,
            "format": ext,
            "registered": True
        }
        
        if exists:
            stat = os.stat(path)
            info["size"] = stat.st_size
            info["modified"] = time.ctime(stat.st_mtime)
        
        return info
    else:
        return {"id": id_name, "registered": False}

def list_registered():
    """Lista todas las hojas registradas."""
    return {id_name: get_sheet_info(id_name) for id_name in LocalSheetBridge._registry.keys()}

def bulk_operations(id_name, operations):
    """Ejecuta múltiples operaciones en una hoja de forma eficiente."""
    sheet = attach(id_name)
    results = []
    
    for op in operations:
        op_type = op.get("type")
        
        if op_type == "write":
            sheet.write(op["cell"], op["value"])
            results.append(f"Written {op['value']} to {op['cell']}")
        
        elif op_type == "read":
            value = sheet.read(op["cell"])
            results.append({"cell": op["cell"], "value": value})
        
        elif op_type == "append":
            sheet.append(op["values"])
            results.append(f"Appended row: {op['values']}")
    
    sheet.save()
    return results

# ============================================================
# Exportación para integración con Orion
# ============================================================

def orion_export():
    """Exporta todas las funciones disponibles para el sistema Orion."""
    return {
        "register": register,
        "attach": attach,
        # NO push/pull aquí
        "status": status,
        "list_synced": list_synced,
        "clear_cache": clear_cache,
        "create_sheet": create_sheet,
        "quick_write": quick_write,
        "quick_read": quick_read,
        "get_sheet_info": get_sheet_info,
        "list_registered": list_registered,
        "bulk_operations": bulk_operations,
        "LocalSheetBridge": LocalSheetBridge,
        "OSyncCache": OSyncCache
    }
